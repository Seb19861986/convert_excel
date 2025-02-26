import pandas as pd
import os
from openpyxl import load_workbook

def convert_excel_to_json(file_path, output_json_path):
    # Charger le fichier Excel avec openpyxl (permet d'accéder aux hyperliens)
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active  # Prend la première feuille

    # Charger les données dans un DataFrame pandas
    df = pd.read_excel(file_path, engine='openpyxl')

    # Supprimer les espaces dans les noms de colonnes
    df.columns = df.columns.str.strip()

    # Extraire les hyperliens de la colonne 'Base légale (1)'
    if 'Base légale (1)' in df.columns:
        hyperlinks = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, 
                                   min_col=df.columns.get_loc('Base légale (1)')+1, 
                                   max_col=df.columns.get_loc('Base légale (1)')+1):
            cell = row[0]
            hyperlinks.append(cell.hyperlink.target if cell.hyperlink else None)
        
        # Ajouter une nouvelle colonne contenant les URL extraites
        df['Base légale (1) - Lien'] = hyperlinks

    # Regroupements de colonnes
    column_groups = {
        "Competence": ['Compétence / responsabilité'],
        "EntiteFocus": ['Entité(s) en focus (contexte)'],
    }

    for new_column, columns in column_groups.items():
        df[new_column] = df[columns].apply(lambda x: ''.join(x.dropna().astype(str)), axis=1)

    # Supprimer les colonnes utilisées pour les regroupements
    columns_to_drop = sum(column_groups.values(), [])
    df_cleaned = df.drop(columns=[col for col in columns_to_drop if col in df.columns])

    # Convertir en JSON
    json_data = df_cleaned.to_json(orient='records', lines=False, force_ascii=False)

    # Sauvegarder en JSON
    with open(output_json_path, 'w', encoding='utf-8') as f:
        f.write(json_data)
